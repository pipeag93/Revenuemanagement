import io
import csv


def extract_text(file_storage) -> str:
    """Extract text from uploaded file — PDF, Excel, CSV or plain text."""
    filename = (file_storage.filename or '').lower()
    data = file_storage.read()

    # PDF
    if filename.endswith('.pdf'):
        return _parse_pdf(data)

    # Excel
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        return _parse_excel(data, filename)

    # CSV
    if filename.endswith('.csv'):
        return _parse_csv(data)

    # Plain text / any other format
    try:
        return data.decode('utf-8', errors='replace')
    except Exception:
        return data.decode('latin-1', errors='replace')


def _parse_pdf(data: bytes) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(data))
        pages = []
        for page in reader.pages[:20]:  # max 20 pages
            text = page.extract_text()
            if text:
                pages.append(text)
        return '\n'.join(pages)
    except Exception as e:
        return f'[Error reading PDF: {e}]'


def _parse_excel(data: bytes, filename: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        lines = []
        for sheet_name in wb.sheetnames[:3]:  # max 3 sheets
            ws = wb[sheet_name]
            lines.append(f'=== Sheet: {sheet_name} ===')
            for row in ws.iter_rows(max_row=200, values_only=True):
                row_vals = [str(v) if v is not None else '' for v in row]
                if any(v.strip() for v in row_vals):
                    lines.append('\t'.join(row_vals))
        return '\n'.join(lines)
    except Exception as e:
        return f'[Error reading Excel: {e}]'


def _parse_csv(data: bytes) -> str:
    try:
        text = data.decode('utf-8', errors='replace')
        reader = csv.reader(io.StringIO(text))
        lines = []
        for i, row in enumerate(reader):
            if i > 500:  # max 500 rows
                break
            lines.append('\t'.join(row))
        return '\n'.join(lines)
    except Exception as e:
        return f'[Error reading CSV: {e}]'
